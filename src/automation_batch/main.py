import requests
import json
from pathlib import Path
import smtplib
from email.message import EmailMessage
import pandas as pd
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.styles import ParagraphStyle
import tkinter as tk
from tkinter import Tk, Button
from tkcalendar import Calendar
from tkinter.ttk import Combobox
from datetime import datetime
from tkinter import ttk
from sku_reader import SKU
from tkinter import font
from openpyxl import load_workbook
import time

def makeEmptyBatch(externalId,batchNotes, max_retries, delay):
    for attempt in range(max_retries):
        try:
            url = "https://api.shipstation.com/v2/batches"

            payload = {
            "external_batch_id": '"'+externalId+'"',
            "batch_notes": '"'+batchNotes+'"'
            }

            headers = {
            "Content-Type": "application/json",
            "api-key": "xFLRkudJ/+cRC+jDUzSiB31fe6E4Vy1m1M1sootREdQ"
            }

            response = requests.post(url, json=payload, headers=headers)
            print(response.status_code)
            response.raise_for_status()
            data = response.json()

            batchNum = data["batch_number"]
            batchId = data["batch_id"]
            return batchNum, batchId
        except requests.HTTPError as e:
            if e.response.status_code == 500:
                print(f"Attempt {attempt+1} failed with 500, retrying in {delay} sec...")
                time.sleep(delay)
            else:
                raise  # re-raise any other HTTP errors
    raise Exception(f"Failed to create empty batch after {max_retries} attempts")

def addShipmentToBatch(batchId,shipmentId):
    url = "https://api.shipstation.com/v2/batches/"+batchId+"/add"

    payload = {
        "shipment_ids": [shipmentId]
    }

    headers = {
    "Content-Type": "application/json",
    "api-key": "xFLRkudJ/+cRC+jDUzSiB31fe6E4Vy1m1M1sootREdQ"
    }

    response = requests.post(url, json=payload, headers=headers)

def changeTags(shipment_id):
    headers = {"api-key": "xFLRkudJ/+cRC+jDUzSiB31fe6E4Vy1m1M1sootREdQ"}
    tag_name = "A) Needs to print"
    url = "https://api.shipstation.com/v2/shipments/"+str(shipment_id)+"/tags/"+tag_name
    response = requests.delete(url, headers=headers)
    tag_name = "B) printed"
    url = "https://api.shipstation.com/v2/shipments/"+str(shipment_id)+"/tags/"+tag_name
    response = requests.post(url, headers=headers)

def makePicklist(df, batchID):
    style = getSampleStyleSheet()
    batch_num = df["batchNum"].iloc[0]
    picklist_folder = Path(__file__).parent / "PickLists"
    picklist_folder.mkdir(exist_ok=True)
    file_path = picklist_folder / f"{batch_num}.pdf"

    pdf = SimpleDocTemplate(str(file_path), pagesize=letter, leftMargin=20, rightMargin=20, topMargin=20)

    small_style = ParagraphStyle(
        name='small',
        fontName='Helvetica',
        fontSize=8
    )

    med_style = ParagraphStyle(
        name='med',
        fontName='Helvetica',
        fontSize=10
    )

    big_style = ParagraphStyle(
        name='big',
        fontName='Helvetica',
        fontSize=12
    )

    batch_style = ParagraphStyle(
        name='batch',
        fontName='Helvetica',
        fontSize=18,
        leading=20
    )

    qty_style = ParagraphStyle(
        name='batch',
        fontName='Helvetica',
        fontSize=12,
        leading=16,
        alignment=2
    )

    data = []
    for _, row in df.iterrows():
        kits = [row["kit1"], row["kit2"], row["kit3"]]
        kits = [str(k) for k in kits if pd.notna(k) and str(k).strip() != ""]
        kit_str = ", ".join(kits)
        
        data.append([
            Paragraph(str(row["sku"]), med_style),
            Paragraph(str(row["title"]), med_style),
            Paragraph(str(row["quantity"]), big_style),
            Paragraph(str(row["binLocation"]), med_style),
            Paragraph(str(kit_str), small_style)
            ])
        
        data.append([Paragraph(str(row["shipment_number"]), small_style)] + ['']*4)

    usable_width = pdf.width
    col_widths = [usable_width*0.2, usable_width*0.45, usable_width*0.05, usable_width*0.10, usable_width*0.2]
    dataTable = Table(data, colWidths=col_widths, hAlign='CENTER')
    header_width = [usable_width*.33, usable_width*.33, usable_width*.33]
    qty_width = [usable_width*1]

    style = TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('BACKGROUND', (0,0), (-1,-1), colors.white)
    ])

    for i in range(len(data)):
        style.add('LINEBELOW', (0,i), (-1,i), 0.5, colors.black)

    for i in range(1, len(data), 2):
        style.add('BACKGROUND', (0,i), (-1,i), colors.whitesmoke)
        style.add('SPAN', (0,i), (-1,i))
        style.add('ALIGN', (0,i), (-1,i), 'CENTER')

    headerStyle = TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('BACKGROUND', (0,0), (-1,-1), colors.white),
        ('LINEBELOW', (0,-1), (-1,1), 2, colors.black)
        ])

    batchStyle = TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('BACKGROUND', (0,0), (-1,-1), colors.white)
        ])

    batch_number = batchID
    batch_notes = df["batchNotes"].iloc[0]
    ship_by_date = df["ship_by_date"].iloc[0]

    batchTable = Table([[
            Paragraph("Batch: "+str(batch_number), batch_style),
            Paragraph(str(batch_notes), batch_style),
            Paragraph("Ship by: "+str(ship_by_date)[5:] if ship_by_date else "", batch_style)
            ]],
        colWidths=header_width, hAlign='CENTER')

    headerTable = Table([[
            Paragraph("SKU", med_style),
            Paragraph("Title", med_style),
            Paragraph("QTY", small_style),
            Paragraph("Bin Location", med_style),
            Paragraph("Kit", med_style)
            ]],
        colWidths=col_widths, hAlign='CENTER'
    )

    qtyStyle = TableStyle([
        ('LINEABOVE', (0,-1), (-1,1), .5, colors.black),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('BACKGROUND', (0,0), (-1,-1), colors.white),
        ('LINEBELOW', (0,-1), (-1,1), 1, colors.black)
        ])

    total_qty = df["quantity"].sum()

    qtyTable = Table(
        [[Paragraph("Total QTY: "+str(total_qty), qty_style)]],
        colWidths=qty_width, hAlign='CENTER')

    batchTable.setStyle(batchStyle)
    headerTable.setStyle(headerStyle)
    dataTable.setStyle(style)
    qtyTable.setStyle(qtyStyle)

    pdf.build([batchTable,headerTable,dataTable,qtyTable])

def tkSelect():
    root = Tk()
    root.update_idletasks()  # calculate sizes
    width = 250
    height = 300
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f"{width}x{height}+{x}+{y}")
    root.title("Select Options")

    combo = Combobox(root, values=["davidk@affinitybands.com", "katharin@affinitybands.com"], state="readonly")
    combo.pack(pady=10)
    combo.current(0)

    cal = Calendar(root, selectmode="day")
    cal.pack(pady=10)

    result = {}

    def submit():
        result['email'] = combo.get()
        selected_date = datetime.strptime(cal.get_date(), "%m/%d/%y")
        result['shipmentDate'] = f"{selected_date.year}-{selected_date.month:02d}-{selected_date.day:02d}"
        root.destroy()

    Button(root, text="OK", command=submit).pack(pady=10)
    root.mainloop()
    return result.get('email'), result.get('shipmentDate')

def getShipments():
    file_path = Path(__file__).parent / "output.json"
    headers = {"api-key": "xFLRkudJ/+cRC+jDUzSiB31fe6E4Vy1m1M1sootREdQ"}
    url = "https://api.shipstation.com/v2/shipments"
    page = 1
    pageSize = 100
    all_shipments = []

    print("starting API pulls")
    print("API call 1")
    params = {
            "shipment_status": "pending",
            "page": page,
            "page_size": pageSize
        }
    
    response = requests.get(url, headers=headers, params=params)
    data = response.json()
    pageNum = data.get("pages")
    print("pageNum = "+str(pageNum))

    while page <= pageNum:
        shipments = data.get("shipments", [])
        if not shipments:
            break
        filtered = [
            s for s in shipments
            if any(tag["name"] == "A) Needs to print" for tag in s.get("tags", []))
        ]
        all_shipments.extend(filtered)
        shipment_ids = [s["shipment_id"] for s in all_shipments]
        shipment_numbers = [s["shipment_number"] for s in all_shipments]
        shipment_dates = [(s["ship_by_date"].split("T")[0] if s["ship_by_date"] else None) for s in all_shipments]

        df_shipments = pd.DataFrame({
            "shipment_id": shipment_ids,
            "shipment_number": shipment_numbers,
            "ship_by_date": shipment_dates
        })
        df_shipments["shipment_number"] = df_shipments["shipment_number"].astype("string")

        page+=1
        if page <= pageNum:
            print("API call "+str(page))
            params = {
                "shipment_status": "pending",
                "page": page,
                "page_size": pageSize
            }
            response = requests.get(url, headers=headers, params=params)
            data = response.json()
    with open(file_path, "w") as f:
        json.dump(data, f, indent=4)
    return df_shipments

def df_nanEmail(df_nan, email):
    sender = "davidk@affinitybands.com"
    app_password = "vrza hjtq xjaa dbor"
    recipient = email

    print('Composing Email')
    msg = EmailMessage()
    msg["Subject"] = "Orders without a Ship By Date"
    msg["From"] = sender
    msg["To"] = recipient
    msg.set_content(df_nan.to_string(index=True))

    print("Sending Email")
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(sender, app_password)
        smtp.send_message(msg)

def getOrder(shipment_id, original_row):
    priorityList = [ 'Priority', '2 Day', '2-Day', '2nd Day', 'Expedited', 'FedEx2DayOneRate', 'FedEx 2 Day Guaranteed Shipping',
        'FedEx Home Delivery', 'FedExStandardOvernight', 'First Class Package International', 'Free FedEx 2 Day over $55',  'Free FedEx 2 Day over $75',
        'Fedex2DayOneRate', 'Fedex 2 Day Guaranteed Shipping', 'Fedex Home Delivery', 'FedexStandardOvernight', 'Free Fedex 2 Day over $55', 
        'Free Fedex 2 Day over $75', 'USPS First Class International', 'USPS Priority Mail 2-3 Day Delivery'
        #'UPS 2nd Day Air' Not actually priority
        ]
    print(shipment_id)
    headers = {"api-key": "xFLRkudJ/+cRC+jDUzSiB31fe6E4Vy1m1M1sootREdQ"}
    url = f"https://api.shipstation.com/v2/shipments/{shipment_id}"
    response = requests.get(url, headers=headers)
    data = response.json()

    rows = []
    for item in data["items"]:
        if item["name"] == "Discounted Item":
            continue
        row = original_row.copy()
        row["sku"] = item["sku"]
        row["quantity"] = item["quantity"]
        row["requested_shipment_service"] = data["requested_shipment_service"]
        sku = SKU(item["sku"])
        if sku.hd == True:
            if sku.prodType == '38':
                row["batchNotes"] = 'HD 38 L+S'
            if sku.prodType == '42':
                row["batchNotes"] = 'HD 42 L+S'
            if sku.prodType == '20' or sku.prodType == '22':
                row["batchNotes"] = 'HD 20+22'
            if sku.productmatch == 'Airpod' or sku.productmatch == 'BudsPro' or sku.productmatch == 'HDXAirpod' or sku.productmatch == 'HDXBudsPro':
                row["batchNotes"] = 'HD Airpods'
            if sku.productmatch == 'Phone':
                row["batchNotes"] = 'Phones'
                if sku.phone == 'AA17' or sku.phone == 'AP17' or sku.phone == 'AM17' or sku.pixelphone == True:
                    row["batchNotes"] = 'iPhone 17'

        if sku.engraved == True or sku.leather == True:
            row["batchNotes"] = 'Engraved+Leather'

        if sku.steel == True or sku.dyesub == True: 
            row["batchNotes"] = 'Steel+Dyesub'

        if sku.screenprint == True or sku.stocked == True or sku.blank == True:
            row["batchNotes"] = 'Screenprint'

        if sku.combo == True:
            row["batchNotes"] = 'Multi'

        if sku.theRest == True:
            row["batchNotes"] = 'theRest'
        if sku.echo == True:
            row["batchNotes"] = 'theRest'
        if data["requested_shipment_service"] in priorityList:
            row["batchNotes"] = "Priority"
        rows.append(row)

    return rows

def edit_batches(df_match, df):
    root = tk.Tk()
    root.update_idletasks()
    width = 1200
    height = 800
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f"{width}x{height}+{x}+{y}")
    root.title("Batch Editor")

    main_frame = tk.Frame(root)
    main_frame.pack(fill="both", expand=True)
    left_frame = tk.Frame(main_frame)
    left_frame.pack(side="left", fill="both", expand=True)
    right_frame = tk.Frame(main_frame)
    right_frame.pack(side="right", fill="both", expand=False)

    cols = ("shipment_number","batchNotes","batchNum","total_qty", "sku_qty")
    tree = ttk.Treeview(left_frame, columns=cols, show="headings")
    for c in cols:
        tree.heading(c, text=c, anchor="w")
        if c == "sku_qty":
            tree.column(c, width=0, stretch=False)  # hide this column
        else:
            tree.column(c, width=140, anchor="w")
    tree.pack(fill="both", expand=True)

    sku_panel = tk.Text(right_frame, width=50, height=20)
    sku_panel.pack(side="right", fill="both", expand=False)
    sku_panel.config(state="disabled")  # make read-only
    for _, row in df.iterrows():
        tree.insert("", "end", values=(
            row["shipment_number"],
            row["batchNotes"],
            row["batchNum"],
            row["total_qty"],
            row["sku_qty"]
        ))
    def show_skus(event):
        selected = tree.selection()
        if not selected:
            return
        row_id = selected[0]
        sku_text = tree.set(row_id, "sku_qty")  # or store in hidden column
        sku_panel.config(state="normal")
        sku_panel.delete("1.0", tk.END)
        sku_panel.insert("1.0", sku_text)
        sku_panel.config(state="disabled")

    tree.bind("<<TreeviewSelect>>", show_skus)
    def on_double_click(event):
        item = tree.identify_row(event.y)
        col = tree.identify_column(event.x)
        if col not in ("#2", "#3"):
            return
        x,y,width,height = tree.bbox(item, col)
        entry = tk.Entry(root)
        entry.place(x=x, y=y+25, width=width)
        if col == "#2":
            entry.insert(0, tree.set(item,"batchNotes"))
        else:
            entry.insert(0, tree.set(item,"batchNum"))
        entry.focus()
        def save(event):
            if col == "#2":
                tree.set(item,"batchNotes",entry.get())
            else:
                tree.set(item,"batchNum",entry.get())
            entry.destroy()
        entry.bind("<Return>", save)
    tree.bind("<Double-1>", on_double_click)
    def done():
        for row_id in tree.get_children():
            vals = tree.item(row_id)["values"]
            shipment = str(vals[0])
            new_notes = str(vals[1])
            new_batch = int(vals[2])
            df_match.loc[
                df_match["shipment_number"] == shipment,
                "batchNotes"
            ] = new_notes
            df_match.loc[
                df_match["shipment_number"] == shipment,
                "batchNum"
            ] = new_batch
        root.destroy()
    btn_font = ("TkDefaultFont", 16, "bold")
    tk.Button(left_frame, text="OK", command=done, font=btn_font,width=8,height=1).pack(pady=5)
    root.mainloop()
    return df_match

def main():
    email, shipmentDate = tkSelect()
    df_shipments = getShipments()

    mask = df_shipments["ship_by_date"] == shipmentDate
    df_match = df_shipments[mask]
    df_match = df_match.reset_index(drop=True)

    if "sku" not in df_match.columns:
        df_match["sku"] = None
    if "quantity" not in df_match.columns:
        df_match["quantity"] = None
    if "requested_shipment_service" not in df_match.columns:
        df_match["requested_shipment_service"] = None
    if "batchNum" not in df_match.columns:
        df_match["batchNum"] = None
    if "batchNotes" not in df_match.columns:
        df_match["batchNotes"] = None
    if "title" not in df_match.columns:
        df_match["title"] = None
    if "binLocation" not in df_match.columns:
        df_match["binLocation"] = None
    if "kit1" not in df_match.columns:
        df_match["kit1"] = None
    if "kit2" not in df_match.columns:
        df_match["kit2"] = None
    if "kit3" not in df_match.columns:
        df_match["kit3"] = None
    
    df_nan = df_shipments[df_shipments["ship_by_date"].isna()].reset_index(drop=True)
    if not df_nan.empty:
        df_nanEmail(df_nan, email)
    else:
        print("No shipments with missing ship_by_date.")

    new_rows = []
    for i, shipment_id in df_match["shipment_id"].items():
        original_row = df_match.loc[i]
        rows = getOrder(shipment_id, original_row)
        new_rows.extend(rows)

    df_match = pd.DataFrame(new_rows)
    df_match = df_match.reset_index(drop=True)

    df_productinfo = pd.read_excel("S:\Workstation - DavidK\!MasterList\Products.xlsx", sheet_name="ProductInfo")
    df_inventory = pd.read_excel("S:\Workstation - DavidK\!MasterList\Inventory.xlsx", sheet_name="KitsBySKU")

    for i, row in df_match.iterrows():
        sku = row["sku"]
        prodMatch = df_productinfo[df_productinfo["SKU"] == sku]
        if not prodMatch.empty:
            df_match.loc[i, "title"] = prodMatch.iloc[0]["Title"]
        invMatch = df_inventory[df_inventory["SKU"] == sku]
        if not invMatch.empty:
            df_match.loc[i, "binLocation"] = invMatch.iloc[0]["WarehouseLocation"]
            df_match.loc[i, "kit1"] = invMatch.iloc[0]["KitComponent1"]
            df_match.loc[i, "kit2"] = invMatch.iloc[0]["KitComponent2"]
            df_match.loc[i, "kit3"] = invMatch.iloc[0]["KitComponent3"]

    for _, group in df_match.groupby("shipment_number"):
        unique_vals = group["batchNotes"].dropna().unique()
        if len(unique_vals) == 1:
            result = unique_vals[0]
        else:
            result = "Multi"
        df_match.loc[group.index, "batchNotes"] = result

    batchNum = pd.read_excel("S:\Workstation - DavidK\Code\Automation_Art\Batch.xlsx", sheet_name="BatchNum", header=None).iloc[0, 0]
    df_match["batchNum"] = df_match.groupby("batchNotes").ngroup() + batchNum
    nextBatch = df_match["batchNum"].max() + 1
    with pd.ExcelWriter("S:\Workstation - DavidK\Code\Automation_Art\Batch.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        pd.DataFrame([[nextBatch, None, "Y", "Y"]]).to_excel(
            writer,
            sheet_name="BatchNum",
            header=False,
            index=False
        )

    df_match["sku_qty"] = df_match["sku"] + " x" + df_match["quantity"].astype(str)
    df_popup = (
        df_match.groupby(["shipment_number","batchNotes","batchNum"], sort=False)
        .agg({
        "sku_qty": lambda x: "\n".join(x),
        "quantity": "sum"
        })
        .rename(columns={"quantity": "total_qty"})
        .reset_index()
    )
    df_match = edit_batches(df_match, df_popup)

    wb = load_workbook("S:\Workstation - DavidK\Code\Automation_Art\Batch.xlsx")
    batches = df_match.groupby("batchNum")
    for batch_number, batch_df in batches:
        print(f"Creating Batch for batch {batch_number}")
        batchNum, batchID = makeEmptyBatch(str(batch_df["batchNum"].iloc[0]),str(batch_df["batchNotes"].iloc[0]),3,2)
        batch_df = batch_df.copy()
        batch_df["batchNum"] = batchNum
        for shipment_id in batch_df["shipment_id"].drop_duplicates():
            addShipmentToBatch(str(batchID), str(shipment_id))
            changeTags(str(shipment_id))
        print(f"Creating PDF for batch {batch_number}")
        batch_df = batch_df.sort_values(
            by=["binLocation", "sku"],
            na_position="last"
        )
        makePicklist(batch_df, batchNum)
        batch_df["batchNum"] = batchID
        sku_list = []
        for _, row in batch_df.iterrows():
            sku_list.extend([row["sku"]] * int(row["quantity"]))
        first = batch_df.iloc[0]
        ship_date = first["ship_by_date"][5:]
        if ship_date not in wb.sheetnames:
            ws = wb.create_sheet(ship_date)
        else:
            ws = wb[ship_date]
        excel_row = [
            first["batchNum"],
            "href",
            "Y",
            "",
            ship_date
        ] + sku_list
        ws.append(excel_row)
    wb.save("S:\Workstation - DavidK\Code\Automation_Art\Batch.xlsx")

    file_path = Path(__file__).parent / "output.json"
    df_match.to_json(file_path, orient="records", indent=4)

    print("done")

main()